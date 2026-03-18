from flask import Flask,render_template,request
from openpyxl import workbook,load_workbook

app = Flask(__name__)

@app.route('/')
def home():
    return render_template("fetchform.html")

@app.route('/submit', methods = ['POST'])
def submit():
    name = request.form['name']
    age = int(request.form['age'])
    gender = request.form['gender']
    agegroup = "Adult" if(age > 18) else "Minor"

    wb = load_workbook("Book1.xlsx")
    sheet = wb.active
    sheet.append([name, age, gender,agegroup])
    wb.save("Book1.xlsx")

    return render_template("fetchform.html", message = f"Information Submitted, Thank You !!!")

@app.route('/last')
def last():
    wb = load_workbook('Book1.xlsx',data_only=True)
    sheet = wb.active
    lastr = sheet.max_row

    nm = sheet.cell(lastr,1).value
    ag = sheet.cell(lastr,2).value
    gn = sheet.cell(lastr,3).value
    aa = sheet.cell(lastr,4).value

    return render_template("fetchform.html", last_row=[nm,ag,gn,aa])

import webbrowser    # to run file auto in browser without cmd
import threading

if __name__ == "__main__":
    threading.Timer(1.5, lambda: webbrowser.open("http://127.0.0.1:5000")).start()
    app.run(debug=False)